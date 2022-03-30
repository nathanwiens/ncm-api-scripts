"""
Cradlepoint NCM API Router Migration Tool
Created by Nathan Wiens (nathan.wiens@cradlepoint.com)

This script will attempt to copy router configuration
from one router to another. The script uses an Excel file of router IDs
to copy from and to.

This script is provided without warranty or liability. Please test thoroughly
before using on any production devices.

Instructions:
    1. Install required Python modules.
        $ pip3 install -r requirements.txt
    2. Update the config.py file with your Account ID and API Keys.
    3. Populate the Excel file with source and destination Router IDs.
        a. Ensure that column 3 is blank for all rows.
    4. Run the script.
        $ python3 router_migration.py

"""

import config
import ncm
import openpyxl
import json

DEBUG = config.DEBUG

def main():
    n = ncm.NcmClient(config.api_keys, log_events=False)
    account_id = config.account_id

    """
    OPEN EXCEL WORKBOOK
    """
    filename = config.EXCEL_FILE
    wb = openpyxl.load_workbook(config.EXCEL_FILE)
    currentSheet = wb.active


    """
    CREATE LIST OF ROUTER SERIAL NUMBERS AND IDS
    """
    routers = {}
    for router in n.get_routers(limit='all'):
        routers[router['serial_number']] = router['id']

    """
    ITERATE THROUGH ALL ROWS IN THE EXCEL FILE
    """
    for rowNum in range(2, currentSheet.max_row + 1):
        src_router_id = str(currentSheet.cell(row=rowNum, column=1).value)
        dst_router_id = str(currentSheet.cell(row=rowNum, column=2).value)
        done = str(currentSheet.cell(row=rowNum, column=3).value)

        if str(done) == "DONE":
            print(f"Router {dst_router_id} already configured. Skipping...")
            continue

        """
        ONLY MAKE CHANGES IF A SERIAL NUMBER IS PRESENT AND THE 'DONE' FIELD IS BLANK.
        THIS PREVENTS RECONFIGURING A SITE THAT THE SCRIPT HAS ALREADY CONFIGURED IN A PREVIOUS RUN.
        """
        if currentSheet.cell(row=rowNum, column=1).value and currentSheet.cell(row=rowNum, column=2).value and done != "DONE":
            print("COPYING CONFIGURATION FOR ROUTER: {} TO ROUTER: {}\n\n".format(src_router_id, dst_router_id))

            src_config = n.get_configuration_managers(router=src_router_id,fields='configuration')[0]

            """
            FOR SOME REASON, THE DHCP LEASE LIFETIME FIELD IS LABELED DIFFERENTLY BETWEEN IBR900 AND IBR1700/R1900
            THIS IS A HACK THAT RENAMES THAT FIELD
            """
            try:
                if src_config['configuration'][0]['lan']['00000000-0d93-319d-8220-4a1fb0372b51']['dhcpd']['lease6_time']:
                    src_config['configuration'][0]['lan']['00000000-0d93-319d-8220-4a1fb0372b51']['dhcpd']['lease6_time'] = src_config['configuration'][0]['lan']['00000000-0d93-319d-8220-4a1fb0372b51']['dhcpd']['valid6_lifetime']
                    del(src_config['configuration'][0]['lan']['00000000-0d93-319d-8220-4a1fb0372b51']['dhcpd']['valid6_lifetime'])
            except KeyError:
                pass

            """
            VPN TUNNEL UUIDS WILL LIKELY BE DIFFERENT BETWEEN SOURCE AND DESTINATION
            THIS IS A HACK THAT MATCHES THEM UP BY STRIPPING THE UNIQUE VALUE
            """
            try:
                for tunnel in list(src_config['configuration'][0]['vpn']['tunnels']):
                    split_string = str(tunnel).split("-", 1)
                    dest_vpn_id = split_string[0][-1]
                    src_config['configuration'][0]['vpn']['tunnels'][dest_vpn_id] = src_config['configuration'][0]['vpn']['tunnels'][tunnel]
                    src_config['configuration'][0]['vpn']['tunnels'][dest_vpn_id].pop('_id_')
                    del(src_config['configuration'][0]['vpn']['tunnels'][tunnel])
            except KeyError:
                pass

            """
            GRE TUNNEL UUIDS WILL LIKELY BE DIFFERENT BETWEEN SOURCE AND DESTINATION
            THIS IS A HACK THAT MATCHES THEM UP BY STRIPPING THE UNIQUE VALUE
            """
            try:
                for tunnel in list(src_config['configuration'][0]['gre']['tunnels']):
                    split_string = str(tunnel).split("-", 1)
                    dest_gre_id = split_string[0][-1]
                    src_config['configuration'][0]['gre']['tunnels'][dest_gre_id] = src_config['configuration'][0]['gre']['tunnels'][tunnel]
                    src_config['configuration'][0]['gre']['tunnels'][dest_gre_id].pop('_id_')
                    del(src_config['configuration'][0]['gre']['tunnels'][tunnel])
            except KeyError:
                pass

            """
            GROUP CONFIGS CAN LEAVE 'SUBTRACTIONS' FROM THE GROUP CONFIG IN THE INDY CONFIG, 
            WHICH WILL THROW ERRORS WHEN PATCHING TO A NEW ROUTER. 
            THIS STRIPS THOSE SUBTRACTIONS FROM THE CONFIG.
            """
            src_config['configuration'][1].clear()


            """
            NCM API CANNOT DECRYPT ENCRYPTED FIELDS LIKE PASSWORDS.
            THIS IS A HACK THAT STRIPS THEM OUT OF THE CONFIG.
            THERE MAY BE MORE FIELDS THAT I'M MISSING. 
            IF YOU GET ERROR 409, CHECK FOR ENCRYPTED FIELDS THAT AREN'T STRIPPED.
            """
            src_config = json.dumps(src_config).replace(', "wpapsk": "*"','').replace('"wpapsk": "*"', '').replace(', "password": "*"', '').replace('"password": "*"', '')
            src_config = json.loads(src_config)


            """
            PUSHING THE FINAL CONFIG TO THE DESTINATION ROUTER
            """
            if DEBUG:
                print(json.dumps(src_config, indent=2))
            result = n.patch_configuration_managers(dst_router_id, src_config)

            if result == "Success":

                """
                SETTING NCM-ONLY FIELDS LIKE LOCATION, CUSTOM1, CUSTOM2
                """
                src_router = n.get_router_by_id(src_router_id, fields='name,custom1,custom2')

                """SET LOCATION FOR GEOVIEW"""
                current_location = n.get_locations(router=src_router_id)
                if current_location:
                    if current_location[0]['method'] == 'manual':
                        n.create_location(account_id, current_location[0]['latitude'], current_location[0]['longitude'], dst_router_id)

                """SET CUSTOM1 and CUSTOM2 FIELDS"""
                n.set_custom1(dst_router_id, src_router['custom1'])
                n.set_custom2(dst_router_id, src_router['custom2'])

                """UPDATE THE EXCEL SHEET WHEN DONE"""
                currentSheet.cell(row=rowNum, column=3).value = "DONE"
                wb.save(filename)
                print("EXCEL SHEET UPDATED AND SAVED.\n")

                print("CONFIGURATION OF ROUTER {} COMPLETE.\n\n\n\n".format(dst_router_id))
            else:

                """UPDATE THE EXCEL SHEET WHEN DONE"""
                currentSheet.cell(row=rowNum, column=3).value = "FAILED"
                wb.save(filename)

                print("CONFIGURATION OF ROUTER {} FAILED.\n\n\n\n".format(dst_router_id))

if __name__ == "__main__":
    main()
